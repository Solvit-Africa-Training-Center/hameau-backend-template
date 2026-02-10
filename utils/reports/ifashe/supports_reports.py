from openpyxl import Workbook
from utils.reports.ifashe.base import BaseExcelReport, BasePDFReport
from programs.models.ifashe_models import (
    SponsoredChild,
    SchoolSupport,
    DressingDistribution,
)


class ChildSupportPDFReport(BasePDFReport):
    def generate(self):
        # self.file_path = file_path

        self.add_title()

        headers = [
            "Child",
            "Family",
            "School Fees",
            "Materials",
            "Total School Cost",
            "Clothes Items Given",
        ]

        rows = []

        for child in SponsoredChild.objects.select_related("family").iterator(200):
            school_support = SchoolSupport.objects.filter(child=child).first()

            clothes_count = DressingDistribution.objects.filter(child=child).count()

            rows.append(
                [
                    child.full_name,
                    child.family.family_name,
                    school_support.school_fees if school_support else 0,
                    school_support.materials_cost if school_support else 0,
                    school_support.total_cost if school_support else 0,
                    clothes_count,
                ]
            )

        self.add_table(headers, rows)

        self.build()


class ChildSupportExcelReport(BaseExcelReport):
    def generate(self, file_path):
        ws = self.wb.create_sheet("Child Support")

        ws.append(
            [
                "Child",
                "Family",
                "School Fees",
                "Materials",
                "Total School Cost",
                "Clothes Items Given",
            ]
        )

        for child in SponsoredChild.objects.select_related("family").iterator(200):
            school_support = SchoolSupport.objects.filter(child=child).first()

            clothes_count = DressingDistribution.objects.filter(child=child).count()

            ws.append(
                [
                    child.full_name,
                    child.family.family_name,
                    school_support.school_fees if school_support else 0,
                    school_support.materials_cost if school_support else 0,
                    school_support.total_cost if school_support else 0,
                    clothes_count,
                ]
            )

        self.wb.save(file_path)


class SupportExcelReport:
    def generate(self, file_path):
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Support Report")

        ws.append(
            [
                "Child",
                "School",
                "Academic Year",
                "Total Cost",
                "Total Paid",
                "Balance",
                "Payment Status",
            ]
        )

        qs = SchoolSupport.objects.select_related("child", "school").prefetch_related(
            "payments"
        )

        for support in qs.iterator(200):
            total_paid = sum(p.amount for p in support.payments.all())

            ws.append(
                [
                    support.child.full_name,
                    support.school.name if support.school else "",
                    support.academic_year,
                    support.total_cost,
                    total_paid,
                    support.total_cost - total_paid,
                    support.payment_status,
                ]
            )

        wb.save(file_path)
